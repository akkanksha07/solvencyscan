"""One-off / re-runnable importer: Data Entry (+ optional Retained Earnings)
tabs of a completed data-collection workbook -> data/raw/<slug>.json.

Usage:
    .venv/bin/python scripts/import_excel.py /path/to/workbook.xlsx

Notes:
- `inventory` and `cash_and_equivalents` are NOT collected by the Data Entry
  tab, so they're written as null -- `quick_ratio` comes out as NaN until
  those are added.
- `retained_earnings` comes from an optional "Retained Earnings" sheet
  (Company/Status/Fiscal Year/Currency/Retained Earnings/Source-Notes/
  Verified columns), matched to Data Entry rows by (company, fiscal year).
  If that sheet is absent, retained_earnings stays null (same as before)
  and Altman Z'' comes out as NaN. The sheet's "Verified" column drives a
  `retained_earnings_note` surfaced in the UI itself (not just here):
    - "PROXY": total equity was used as a stand-in for the true retained-
      earnings sub-line -- this makes Z'' slightly OPTIMISTIC, since equity
      also includes share capital/premium/other reserves.
    - Source/Notes text containing "approx": the figure is a calculated
      estimate, not read directly off a balance sheet.
    - A negative value: flagged as an accumulated deficit, which is a
      normal capital-return pattern for a mature company (buybacks/
      dividends > cumulative net income), not automatically a distress sign.
- `total_liabilities` isn't a workbook column either; it's derived as
  Total Assets - Equity (standard accounting identity), which is actually
  more robust than the workbook's own "Total Debt" column since the latter
  explicitly excludes lease liabilities for several companies (see their
  Source Page / Note text) while the identity captures everything.
- prior_total_assets/prior_total_equity are chained from each company's own
  earlier analysed year (not from the workbook), to average ROA/ROE/asset-
  turnover denominators for the 2nd and 3rd year of each company's window.
- Duplicate (company, fiscal year) rows are resolved by preferring the one
  marked Verified=Y; if that's ambiguous the conflict is printed and BOTH
  are skipped so a bad row can't silently overwrite a good one.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent
DATA_RAW_DIR = REPO_ROOT / "data" / "raw"

NAME_TO_SLUG = {
    "Carillion": "carillion",
    "Thomas Cook": "thomas_cook",
    "Wirecard": "wirecard",
    "Enron": "enron",
    "Apple": "apple",
    "Walmart": "walmart",
    "Microsoft": "microsoft",
    "Unilever": "unilever",
}


def load_rows(xlsx_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Data Entry"]
    all_rows = list(ws.iter_rows(values_only=True))
    header = all_rows[2]
    rows = []
    for raw in all_rows[3:]:
        if raw[0] is None:
            continue
        d = dict(zip(header, raw))
        company = (d.get("Company") or "").strip()
        if company not in NAME_TO_SLUG:
            continue  # skips the "EXAMPLE ..." placeholder row too
        rows.append(d)
    return rows


def _proxy_note(company: str) -> str:
    return (
        f"Retained earnings for {company} uses total shareholders' equity as a "
        f"proxy -- the actual retained-earnings sub-line wasn't separately "
        f"extractable. This makes the Altman Z''-Score slightly OPTIMISTIC "
        f"relative to the true figure, since total equity also includes share "
        f"capital, share premium and other reserves and so typically exceeds "
        f"retained earnings. Disclosed here as a limitation, not corrected "
        f"silently."
    )


def _estimate_note(company: str, fy: str) -> str:
    return (
        f"{company}'s {fy} retained earnings figure is a calculated estimate "
        f"(opening balance + net income - dividends - buybacks), not a number "
        f"read directly off a published balance sheet like its other years."
    )


def _negative_note(company: str) -> str:
    return (
        f"{company}'s retained earnings is negative (an accumulated "
        f"deficit) in this year. This can be a normal result of returning "
        f"capital to shareholders (dividends/buybacks) faster than "
        f"cumulative net income -- common for mature, cash-generative "
        f"companies -- not automatically a distress signal, though it does "
        f"shrink the equity cushion the Z''-Score's X4 term relies on."
    )


def load_retained_earnings(xlsx_path: Path) -> dict[tuple[str, str], dict]:
    """Optional "Retained Earnings" sheet -> {(company, fiscal_year): {value, note}}."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "Retained Earnings" not in wb.sheetnames:
        return {}
    ws = wb["Retained Earnings"]
    all_rows = list(ws.iter_rows(values_only=True))
    header = all_rows[0]
    out: dict[tuple[str, str], dict] = {}
    for raw in all_rows[1:]:
        if raw[0] is None or (raw[1] is None and raw[2] is None):
            continue  # blank row or the "NOTE:" banner row
        d = dict(zip(header, raw))
        company = (d.get("Company") or "").strip()
        if company not in NAME_TO_SLUG:
            continue
        fy = str(d["Fiscal Year"])
        value = d.get("Retained Earnings")
        if value is None:
            continue
        value = float(value)
        verified = str(d.get("Verified (Y/N)") or "").strip().upper()
        source_note = str(d.get("Source / Notes") or "")

        note = None
        if verified == "PROXY":
            note = _proxy_note(company)
        elif "approx" in source_note.lower():
            note = _estimate_note(company, f"FY{fy}")
        elif value < 0:
            note = _negative_note(company)

        out[(company, fy)] = {"value": value, "note": note, "verified": verified}
    return out


def dedupe(rows: list[dict]) -> tuple[list[dict], list[str]]:
    """Collapse duplicate (company, fiscal year) rows, preferring Verified=Y.
    Returns (clean_rows, warning_messages)."""
    groups: dict[tuple[str, str], list[dict]] = {}
    for d in rows:
        key = (d["Company"], str(d["Fiscal Year"]))
        groups.setdefault(key, []).append(d)

    clean, warnings = [], []
    for (company, fy), group in groups.items():
        if len(group) == 1:
            clean.append(group[0])
            continue
        verified_y = [d for d in group if d.get("Verified (Y/N)") == "Y"]
        if len(verified_y) == 1:
            clean.append(verified_y[0])
            warnings.append(
                f"DUPLICATE ROW: {company} {fy} appeared {len(group)}x -- kept the "
                f"Verified=Y version, discarded the other(s) "
                f"(discarded had Interest Expense={[d.get('Interest Expense') for d in group if d is not verified_y[0]]})."
            )
        else:
            warnings.append(
                f"UNRESOLVED DUPLICATE: {company} {fy} appeared {len(group)}x with "
                f"ambiguous Verified status -- SKIPPED both, needs manual fix in the workbook."
            )
    return clean, warnings


def to_raw_year(d: dict, prior: dict | None, re_entry: dict | None) -> dict:
    fy = str(d["Fiscal Year"])
    total_assets = float(d["Total Assets"])
    total_equity = float(d["Equity"])
    cogs = d.get("COGS")
    source = f"{d.get('Source URL', '')} | {d.get('Source Page / Note', '')}".strip(" |")
    return {
        "fiscal_year": f"FY{fy}",
        "total_assets": total_assets,
        "total_liabilities": total_assets - total_equity,
        "total_equity": total_equity,
        "current_assets": float(d["Current Assets"]),
        "current_liabilities": float(d["Current Liabilities"]),
        "receivables": float(d["Receivables"]),
        "inventory": None,
        "retained_earnings": re_entry["value"] if re_entry else None,
        "retained_earnings_note": re_entry["note"] if re_entry else None,
        "cash_and_equivalents": None,
        "revenue": float(d["Revenue"]),
        "cogs": float(cogs) if cogs is not None else None,
        "operating_income": float(d["EBIT"]),
        "interest_expense": float(d["Interest Expense"]),
        "net_income": float(d["Net Income"]),
        "cfo": float(d["Operating Cash Flow"]),
        "capex": float(d["Capex"]),
        "prior_total_assets": float(prior["Total Assets"]) if prior else None,
        "prior_total_equity": float(prior["Equity"]) if prior else None,
        "source": f"[{d.get('Currency', '?')}] {source}",
    }


def main() -> None:
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else None
    if not xlsx_path or not xlsx_path.exists():
        print("Usage: import_excel.py /path/to/workbook.xlsx")
        sys.exit(1)

    rows = load_rows(xlsx_path)
    clean, warnings = dedupe(rows)
    re_lookup = load_retained_earnings(xlsx_path)

    by_slug: dict[str, list[dict]] = {}
    for d in clean:
        slug = NAME_TO_SLUG[d["Company"]]
        by_slug.setdefault(slug, []).append(d)

    unverified, re_missing, re_caveats = [], [], []
    for slug, company_rows in by_slug.items():
        company_rows.sort(key=lambda d: str(d["Fiscal Year"]))
        years = []
        prior = None
        for d in company_rows:
            company, fy = d["Company"], str(d["Fiscal Year"])
            re_entry = re_lookup.get((company, fy))
            years.append(to_raw_year(d, prior, re_entry))
            prior = d
            if d.get("Verified (Y/N)") != "Y":
                unverified.append(f"{company} FY{fy}")
            if re_entry is None:
                re_missing.append(f"{company} FY{fy}")
            elif re_entry["note"]:
                re_caveats.append(f"{company} FY{fy}: {re_entry['note']}")
        out_path = DATA_RAW_DIR / f"{slug}.json"
        out_path.write_text(json.dumps({"years": years}, indent=2) + "\n")
        print(f"wrote {out_path.relative_to(REPO_ROOT)} ({len(years)} years)")

    print()
    if warnings:
        print("=== Duplicate-row warnings ===")
        for w in warnings:
            print(f"  - {w}")
    if unverified:
        print("=== Verified: N rows (double-check these) ===")
        for u in unverified:
            print(f"  - {u}")
    if re_missing:
        print("=== No retained earnings found (Altman Z'' will show n/a) ===")
        for m in re_missing:
            print(f"  - {m}")
    if re_caveats:
        print("=== Retained earnings caveats (also shown in the UI) ===")
        for c in re_caveats:
            print(f"  - {c}")
    missing = set(NAME_TO_SLUG) - {d["Company"] for d in clean}
    if missing:
        print("=== Companies with NO usable rows ===")
        for m in missing:
            print(f"  - {m}")


if __name__ == "__main__":
    main()
